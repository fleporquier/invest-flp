"""Render scan results to console, Markdown, Excel and JSON.

A *result* is a flat dict combining drop metrics, technical indicators, the
Claude verdict and PEA / Trade Republic eligibility for one ticker. Heavy
optional dependencies (``rich``, ``openpyxl``) are imported lazily so the
module stays importable without them.

The JSON output is **automatically redacted** via :mod:`dashboard.redact`
because it is intended to feed the public dashboard.
"""

from __future__ import annotations

import json
import logging
from datetime import date as date_cls, datetime, timezone
from pathlib import Path

logger = logging.getLogger(__name__)

_VERDICT_ORDER = ["ACHETER", "ATTENDRE", "EVITER"]
_VERDICT_COLOR = {"ACHETER": "green", "ATTENDRE": "yellow", "EVITER": "red"}
_VERDICT_EMOJI = {"ACHETER": "🟢", "ATTENDRE": "🟡", "EVITER": "🔴"}

_SHEET_BY_CATEGORY = {
    "us_tech_ai": "US_Tech_IA",
    "china_tech": "China_Tech",
    "etf": "ETF",
}


def _stars(conviction: object) -> str:
    """Render a 1-10 conviction as a five-star string.

    Args:
        conviction: Conviction value (int-like) or ``None``.

    Returns:
        A string of filled/empty stars, or ``"-"`` if unknown.
    """
    try:
        value = int(conviction)  # type: ignore[arg-type]
    except (TypeError, ValueError):
        return "-"
    filled = max(0, min(5, round(value / 2)))
    return "★" * filled + "☆" * (5 - filled)


def render_console(results: list[dict], top: int = 10) -> None:
    """Print a coloured summary table of the top results.

    Args:
        results: Scan results, assumed already ranked.
        top: Number of rows to display.
    """
    try:
        from rich.console import Console
        from rich.table import Table
    except ImportError:  # pragma: no cover - dependency guard
        logger.warning("rich not installed; falling back to plain output")
        for res in results[:top]:
            logger.info(
                "%s %s conviction=%s score=%s",
                res.get("ticker"),
                res.get("verdict"),
                res.get("conviction"),
                res.get("technical_score"),
            )
        return

    table = Table(title=f"Opportunities — top {min(top, len(results))}")
    for col in ("Ticker", "Verdict", "Conviction", "Tech", "Baisse 5j", "Baisse 1m", "PEA/TR", "Entrée"):
        table.add_column(col)

    for res in results[:top]:
        verdict = str(res.get("verdict", "?"))
        color = _VERDICT_COLOR.get(verdict, "white")
        pea_tr = _format_eligibility(res)
        table.add_row(
            str(res.get("ticker", "")),
            f"[{color}]{verdict}[/{color}]",
            _stars(res.get("conviction")),
            str(res.get("technical_score", "")),
            _fmt_pct(res.get("drop_5d")),
            _fmt_pct(res.get("drop_1m")),
            pea_tr,
            str(res.get("niveau_entree_suggere") or "-"),
        )
    Console().print(table)


def _format_eligibility(res: dict) -> str:
    """Render a compact PEA / TR / ETF-alt badge.

    Args:
        res: A single result dict.

    Returns:
        Short eligibility string, e.g. ``"PEA"``, ``"TR"`` or ``"→PUST.PA"``.
    """
    if res.get("pea"):
        return "PEA"
    parts: list[str] = []
    if res.get("tr"):
        parts.append("TR")
    if res.get("etf_pea_alt"):
        parts.append(f"→{res['etf_pea_alt']}")
    return " ".join(parts) if parts else "-"


def _fmt_pct(value: object) -> str:
    """Format a percentage value for display.

    Args:
        value: A number or ``None``.

    Returns:
        Formatted percentage string, or ``"n/a"``.
    """
    if value is None:
        return "n/a"
    try:
        return f"{float(value):+.1f}%"
    except (TypeError, ValueError):
        return "n/a"


def write_markdown(results: list[dict], output_dir: str | Path, run_date: date_cls | None = None) -> Path:
    """Write a dated Markdown report grouped by verdict.

    Args:
        results: Scan results.
        output_dir: Destination directory.
        run_date: Report date (defaults to today).

    Returns:
        Path to the written Markdown file.
    """
    run_date = run_date or date_cls.today()
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"opportunities_{run_date.isoformat()}.md"

    lines: list[str] = [
        f"# Opportunities Scanner — {run_date.isoformat()}",
        "",
        f"{len(results)} candidat(s) qualifié(s). Focus tech US / Chine / IA.",
        "",
        "> ⚠️ Données de marché via yfinance ; signaux argumentés, non garantis. "
        "La décision et l'exécution restent à l'utilisateur.",
        "",
    ]

    by_verdict: dict[str, list[dict]] = {v: [] for v in _VERDICT_ORDER}
    for res in results:
        by_verdict.setdefault(str(res.get("verdict", "ATTENDRE")), []).append(res)

    for verdict in _VERDICT_ORDER:
        bucket = by_verdict.get(verdict, [])
        if not bucket:
            continue
        emoji = _VERDICT_EMOJI.get(verdict, "")
        lines.append(f"## {emoji} {verdict} ({len(bucket)})")
        lines.append("")
        lines.append("| Ticker | Nom | Baisse 5j | Baisse 1m | RSI | Tech | Conv. | Type baisse | Entrée | PEA/TR/ETF |")
        lines.append("|---|---|---|---|---|---|---|---|---|---|")
        for res in bucket:
            lines.append(
                "| {ticker} | {name} | {d5} | {d1m} | {rsi} | {tech} | {conv} | {tb} | {entry} | {elig} |".format(
                    ticker=res.get("ticker", ""),
                    name=res.get("name", ""),
                    d5=_fmt_pct(res.get("drop_5d")),
                    d1m=_fmt_pct(res.get("drop_1m")),
                    rsi=res.get("rsi_14", "n/a"),
                    tech=res.get("technical_score", ""),
                    conv=res.get("conviction", ""),
                    tb=res.get("type_baisse", ""),
                    entry=res.get("niveau_entree_suggere") or "-",
                    elig=_format_eligibility(res),
                )
            )
        lines.append("")
        for res in bucket:
            reason = res.get("raison_baisse")
            catalyst = res.get("catalyseur_rebond")
            if reason or catalyst:
                lines.append(f"- **{res.get('ticker')}** — {reason or ''}"
                             + (f" _Catalyseur :_ {catalyst}" if catalyst else ""))
        lines.append("")

    path.write_text("\n".join(lines), encoding="utf-8")
    logger.info("Markdown report written to %s", path)
    return path


def write_excel(results: list[dict], output_dir: str | Path, run_date: date_cls | None = None) -> Path:
    """Write a dated Excel report with one sheet per category.

    Args:
        results: Scan results.
        output_dir: Destination directory.
        run_date: Report date (defaults to today).

    Returns:
        Path to the written workbook.

    Raises:
        ImportError: If ``openpyxl`` is not installed.
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError as exc:  # pragma: no cover - dependency guard
        raise ImportError("openpyxl is required for Excel output") from exc

    run_date = run_date or date_cls.today()
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"opportunities_{run_date.isoformat()}.xlsx"

    headers = [
        "Ticker", "Nom", "Dernier", "Baisse 5j %", "Baisse 1m %", "RSI 14",
        "MACD", "Range 52w", "Score tech", "Verdict", "Type baisse",
        "Conviction", "Horizon", "Entrée", "PEA", "TR", "ETF PEA alt",
    ]
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F4E78")

    workbook = Workbook()
    workbook.remove(workbook.active)

    by_sheet: dict[str, list[dict]] = {name: [] for name in _SHEET_BY_CATEGORY.values()}
    for res in results:
        sheet = _SHEET_BY_CATEGORY.get(str(res.get("category")), "US_Tech_IA")
        by_sheet[sheet].append(res)

    for sheet_name, rows in by_sheet.items():
        worksheet = workbook.create_sheet(title=sheet_name)
        worksheet.append(headers)
        for col_idx in range(1, len(headers) + 1):
            cell = worksheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
        for res in rows:
            worksheet.append(
                [
                    res.get("ticker", ""),
                    res.get("name", ""),
                    res.get("last_price"),
                    res.get("drop_5d"),
                    res.get("drop_1m"),
                    res.get("rsi_14"),
                    res.get("macd_signal"),
                    res.get("range_52w_position"),
                    res.get("technical_score"),
                    res.get("verdict"),
                    res.get("type_baisse"),
                    res.get("conviction"),
                    res.get("horizon"),
                    res.get("niveau_entree_suggere"),
                    "Oui" if res.get("pea") else "Non",
                    "Oui" if res.get("tr") else "Non",
                    res.get("etf_pea_alt") or "",
                ]
            )
        _autosize(worksheet, len(headers))

    workbook.save(path)
    logger.info("Excel report written to %s", path)
    return path


def _autosize(worksheet, n_cols: int) -> None:
    """Roughly auto-size worksheet columns to their content.

    Args:
        worksheet: The openpyxl worksheet.
        n_cols: Number of columns to size.
    """
    from openpyxl.utils import get_column_letter

    for col_idx in range(1, n_cols + 1):
        letter = get_column_letter(col_idx)
        longest = max(
            (len(str(cell.value)) for cell in worksheet[letter] if cell.value is not None),
            default=8,
        )
        worksheet.column_dimensions[letter].width = min(max(longest + 2, 10), 45)


def write_json(
    results: list[dict],
    output_dir: str | Path,
    run_date: date_cls | None = None,
    redact: bool = True,
) -> Path:
    """Write a redacted JSON snapshot fit for the public dashboard.

    Args:
        results: Scan results.
        output_dir: Destination directory.
        run_date: Report date (defaults to today).
        redact: When ``True`` (default), euro amounts, market caps and
            similar absolute figures are stripped before serialisation.

    Returns:
        Path to the written JSON file.
    """
    run_date = run_date or date_cls.today()
    out_dir = Path(output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)
    path = out_dir / f"opportunities_{run_date.isoformat()}.json"

    payload = {
        "generated_at": datetime.now(tz=timezone.utc).isoformat(timespec="seconds"),
        "results": results,
    }
    if redact:
        from dashboard.redact import redact as _redact

        payload = _redact(payload)
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    logger.info("JSON report written to %s", path)
    return path


def generate_reports(
    results: list[dict],
    formats: list[str],
    output_dir: str | Path,
    top: int = 10,
    run_date: date_cls | None = None,
) -> dict[str, Path | None]:
    """Generate every requested report format.

    Args:
        results: Scan results.
        formats: Subset of ``{"console", "markdown", "excel", "json"}``.
        output_dir: Destination directory for file outputs.
        top: Number of rows for the console table.
        run_date: Report date (defaults to today).

    Returns:
        Mapping of format name to the written path (``None`` for console).
    """
    written: dict[str, Path | None] = {}
    if "console" in formats:
        render_console(results, top=top)
        written["console"] = None
    if "markdown" in formats:
        written["markdown"] = write_markdown(results, output_dir, run_date)
    if "excel" in formats:
        written["excel"] = write_excel(results, output_dir, run_date)
    if "json" in formats:
        written["json"] = write_json(results, output_dir, run_date)
    return written
